"""
Create Excel submission file for India Predicts 2026 competition
Converts final_submission_2026.csv to the required Excel format
"""

import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def create_submission_excel():
    """Create formatted Excel submission file."""
    
    logger.info("\n" + "="*70)
    logger.info("CREATING EXCEL SUBMISSION FILE")
    logger.info("="*70 + "\n")
    
    # Load data
    logger.info("Loading prediction data...")
    df = pd.read_csv('outputs/final_submission_2026.csv')
    logger.info(f"✓ Loaded {len(df)} predictions\n")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    
    # Add header row
    headers = ['State', 'Constituency', 'Predicted Winner']
    ws.append(headers)
    
    # Format header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data rows
    logger.info("Adding data rows...")
    for idx, row in df.iterrows():
        ws.append([
            row['state'],
            row['constituency'],
            row['predicted_winner']
        ])
    
    # Format data rows
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # State
    ws.column_dimensions['B'].width = 25  # Constituency
    ws.column_dimensions['C'].width = 30  # Predicted Winner
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    logger.info("Adding summary sheet...")
    summary = wb.create_sheet("Summary")
    
    summary['A1'] = "India Predicts 2026 - Election Prediction Challenge"
    summary['A1'].font = Font(bold=True, size=14, color="1F4E78")
    
    summary['A3'] = "Submission Summary"
    summary['A3'].font = Font(bold=True, size=12)
    
    row_num = 4
    summary[f'A{row_num}'] = "Total Predictions:"
    summary[f'B{row_num}'] = len(df)
    row_num += 1
    
    summary[f'A{row_num}'] = "Submission Date:"
    summary[f'B{row_num}'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
    row_num += 2
    
    summary[f'A{row_num}'] = "State-wise Breakdown:"
    summary[f'A{row_num}'].font = Font(bold=True, size=11)
    row_num += 1
    
    state_counts = df['state'].value_counts().sort_index()
    for state, count in state_counts.items():
        summary[f'A{row_num}'] = state
        summary[f'B{row_num}'] = count
        row_num += 1
    
    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 15
    
    # Save
    output_file = 'India_Predicts_2026_SUBMISSION.xlsx'
    wb.save(output_file)
    
    logger.info(f"\n{'='*70}")
    logger.info(f"✓ SUBMISSION FILE CREATED")
    logger.info(f"{'='*70}\n")
    logger.info(f"File: {output_file}")
    logger.info(f"Size: {len(df)} constituencies")
    logger.info(f"Format: Excel (.xlsx)")
    logger.info(f"Status: READY TO SUBMIT\n")
    
    # Verification
    logger.info(f"{'='*70}")
    logger.info("VERIFICATION")
    logger.info(f"{'='*70}\n")
    
    logger.info("State distribution:")
    for state, count in state_counts.items():
        logger.info(f"  {state}: {count}")
    
    logger.info(f"\nSample predictions:")
    for idx, row in df.head(5).iterrows():
        logger.info(f"  {row['state']:15s} | {row['constituency']:25s} | {row['predicted_winner']}")
    
    logger.info(f"\nFinal rows:")
    for idx, row in df.tail(3).iterrows():
        logger.info(f"  {row['state']:15s} | {row['constituency']:25s} | {row['predicted_winner']}")
    
    logger.info(f"\n{'='*70}")
    logger.info("NEXT STEPS")
    logger.info(f"{'='*70}\n")
    
    logger.info("1. ✓ Excel file ready: India_Predicts_2026_SUBMISSION.xlsx")
    logger.info("2. Create Methodology.pdf or Methodology.docx (150+ words)")
    logger.info("3. Register at: https://forms.gle/2kunspu9A7txkngQ9")
    logger.info("4. Submit both files before April 30, 2026, 11:59 PM")
    logger.info("5. Verify accuracy on May 4, 2026\n")


def create_methodology_template():
    """Create a methodology template file."""
    
    template = """METHODOLOGY NOTE - INDIA PREDICTS 2026
======================================

Your Name/Team Name: [ENTER YOUR NAME OR TEAM NAME]
Submission Date: [DATE]

1. DATA SOURCES
===============
List all data sources you used:
- Myneta.info: Candidate information (names, parties, criminal cases, education, assets)
- Election Commission of India (ECI): Official constituency and state information
- [Add other sources you used]

2. DATA COLLECTION & PREPROCESSING
==================================
Describe how you collected and prepared your data:
- Downloaded candidate data from Myneta for all 5 states
- Extracted constituency names and party affiliations
- Cleaned candidate names (removed titles, standardized formats)
- Merged with historical election data (if used)

3. PREDICTION METHODOLOGY
=========================
Explain your prediction approach (min 150 words):

My model predicts winners by analyzing candidate characteristics and party strength.
Key factors considered:

a) Candidate Strength:
   - Educational background
   - Professional experience
   - Criminal record status
   - Asset accumulation

b) Party Factors:
   - Historical performance in state/constituency
   - Party infrastructure
   - Alliance patterns

c) Constituency Factors:
   - [Describe any constituency-specific analysis]

Prediction Logic:
[Describe how you combined these factors to make predictions]

4. MODEL ASSUMPTIONS
====================
- Assumption 1: [Your assumption]
- Assumption 2: [Your assumption]
- [Add more assumptions]

5. STRENGTHS OF THIS APPROACH
=============================
- Uses official Myneta data sourced directly from candidates
- Incorporates multiple factors beyond just party
- [Other strengths]

6. LIMITATIONS & CHALLENGES
===========================
- Some new candidates lack historical data
- Election dynamics can be unpredictable
- Regional/linguistic factors not fully captured
- [Other limitations]

7. CONFIDENCE LEVEL
===================
Overall prediction confidence: [HIGH / MEDIUM / LOW]
States where I'm most confident: [STATE NAMES]
States where I'm less confident: [STATE NAMES]
Reasoning: [Explain confidence levels]

---
Total Words: [COUNT YOUR WORDS - MUST BE 150+]
"""
    
    with open('METHODOLOGY_TEMPLATE.txt', 'w') as f:
        f.write(template)
    
    logger.info("\n✓ Created METHODOLOGY_TEMPLATE.txt")
    logger.info("Edit this file to create your methodology note")


if __name__ == '__main__':
    create_submission_excel()
    create_methodology_template()
