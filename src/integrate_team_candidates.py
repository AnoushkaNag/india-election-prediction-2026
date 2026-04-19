"""
Integration System: Merge Team's Candidate Data with Predictions
Accepts Excel file with candidate names and integrates into final submission
"""

import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from difflib import SequenceMatcher
import os

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def normalize_string(s):
    """Normalize string for fuzzy matching."""
    import re
    return re.sub(r'[^a-z0-9\s]', '', str(s).lower().strip())


def fuzzy_match(s1, s2, threshold=0.85):
    """Fuzzy string matching for constituency names."""
    s1_norm = normalize_string(s1)
    s2_norm = normalize_string(s2)
    ratio = SequenceMatcher(None, s1_norm, s2_norm).ratio()
    return ratio >= threshold


def load_team_candidate_data(file_path):
    """Load candidate data from team's Excel file."""
    logger.info(f"\nLoading team data from: {file_path}")
    
    if not os.path.exists(file_path):
        logger.error(f"✗ File not found: {file_path}")
        return None
    
    try:
        df = pd.read_csv(file_path) if file_path.endswith('.csv') else pd.read_excel(file_path)
        logger.info(f"✓ Loaded {len(df)} candidates")
        logger.info(f"  Columns: {list(df.columns)}")
        return df
    except Exception as e:
        logger.error(f"✗ Error loading file: {e}")
        return None


def integrate_team_candidates():
    """
    Main integration:
    1. Load predictions (824 constituencies)
    2. Load team's candidate data
    3. Match by State + Constituency
    4. Create final submission
    """
    
    logger.info("\n" + "="*70)
    logger.info("TEAM CANDIDATE DATA INTEGRATION")
    logger.info("="*70 + "\n")
    
    # Load predictions
    logger.info("Loading predictions...")
    predictions = pd.read_csv('outputs/final_submission.csv')
    logger.info(f"✓ Loaded {len(predictions)} predictions\n")
    
    # Try to load team data
    candidate_files = [
        'Candidate_Names_From_Myneta.xlsx',
        'Candidate_Names_From_Myneta.csv',
        'candidate_data.xlsx',
        'candidate_data.csv',
    ]
    
    team_data = None
    loaded_file = None
    
    for file in candidate_files:
        if os.path.exists(file):
            team_data = load_team_candidate_data(file)
            loaded_file = file
            if team_data is not None:
                break
    
    if team_data is None:
        logger.error("\n✗ No team candidate file found!")
        logger.info("Expected file names:")
        for f in candidate_files:
            logger.info(f"  - {f}")
        return None
    
    logger.info(f"✓ Using file: {loaded_file}\n")
    
    # Standardize column names
    team_data.columns = [col.lower().strip() for col in team_data.columns]
    
    # Identify columns
    state_col = next((col for col in team_data.columns if 'state' in col), None)
    name_col = next((col for col in team_data.columns if 'candidate' in col or 'name' in col), None)
    const_col = next((col for col in team_data.columns if 'constituency' in col or 'const' in col), None)
    
    if not all([state_col, name_col, const_col]):
        logger.error(f"✗ Could not find required columns!")
        logger.error(f"  State column: {state_col}")
        logger.error(f"  Name column: {name_col}")
        logger.error(f"  Constituency column: {const_col}")
        return None
    
    logger.info(f"✓ Detected columns:")
    logger.info(f"  State: {state_col}")
    logger.info(f"  Candidate: {name_col}")
    logger.info(f"  Constituency: {const_col}\n")
    
    # Create lookup
    logger.info("Building candidate lookup...")
    candidate_lookup = {}
    
    for idx, row in team_data.iterrows():
        state = str(row[state_col]).strip()
        name = str(row[name_col]).strip()
        const = str(row[const_col]).strip()
        
        key = f"{state}|{const}".lower()
        candidate_lookup[key] = name
    
    logger.info(f"✓ Indexed {len(candidate_lookup)} candidates\n")
    
    # Merge
    logger.info("Merging predictions with candidate data...")
    merged = []
    matched = 0
    unmatched = []
    
    for idx, pred_row in predictions.iterrows():
        state = pred_row['state'].strip()
        const = pred_row['constituency'].strip()
        pred_party = pred_row['predicted_winner'].strip()
        
        # Try exact match
        key = f"{state}|{const}".lower()
        
        if key in candidate_lookup:
            winner = candidate_lookup[key]
            matched += 1
        else:
            # Try fuzzy match
            fuzzy_match_found = False
            for lookup_key, cand_name in candidate_lookup.items():
                lookup_state, lookup_const = lookup_key.split('|')
                if (normalize_string(lookup_state) == normalize_string(state) and
                    fuzzy_match(lookup_const, const, 0.8)):
                    winner = cand_name
                    matched += 1
                    fuzzy_match_found = True
                    break
            
            if not fuzzy_match_found:
                winner = pred_party  # Fallback to party
                unmatched.append({
                    'state': state,
                    'constituency': const,
                    'reason': 'No match in team data'
                })
        
        merged.append({
            'state': state,
            'constituency': const,
            'predicted_winner': winner
        })
    
    merged_df = pd.DataFrame(merged)
    
    # Report
    logger.info(f"\n{'='*70}")
    logger.info("MERGE RESULTS")
    logger.info(f"{'='*70}\n")
    
    party_codes = ['INC', 'BJP', 'AITC', 'DMK', 'AIUDF', 'AIADMK', 'BJD', 'UPPL', 'BRS', 'SP', 'BSP', 'AAP']
    actual_names = merged_df[~merged_df['predicted_winner'].isin(party_codes)]
    
    logger.info(f"Total predictions: {len(merged_df)}")
    logger.info(f"  ✓ Matched to candidate names: {matched} ({100*matched//len(merged_df)}%)")
    logger.info(f"  ✗ Fell back to party codes: {len(merged_df)-matched} ({100*(len(merged_df)-matched)//len(merged_df)}%)")
    
    logger.info(f"\nData composition:")
    logger.info(f"  With actual names: {len(actual_names):5d} ({100*len(actual_names)//len(merged_df)}%)")
    logger.info(f"  With party codes:  {len(merged_df)-len(actual_names):5d} ({100*(len(merged_df)-len(actual_names))//len(merged_df)}%)")
    
    # Validation
    logger.info(f"\n{'='*70}")
    logger.info("VALIDATION")
    logger.info(f"{'='*70}\n")
    
    logger.info(f"✓ Total rows: {len(merged_df)} (expected 824)")
    logger.info(f"✓ No missing values: {merged_df['predicted_winner'].isnull().sum() == 0}")
    logger.info(f"✓ No duplicates: {len(merged_df) == len(merged_df.drop_duplicates())}")
    
    logger.info(f"\nState distribution:")
    for state, count in merged_df['state'].value_counts().sort_index().items():
        logger.info(f"  {state:20s}: {count:5d}")
    
    # Save CSV
    csv_path = 'outputs/final_submission_2026.csv'
    merged_df[['state', 'constituency', 'predicted_winner']].to_csv(csv_path, index=False)
    logger.info(f"\n✓ Saved to {csv_path}")
    
    # Create Excel
    create_excel_submission(merged_df)
    
    return merged_df


def create_excel_submission(merged_df):
    """Create formatted Excel submission file."""
    
    logger.info("\nCreating Excel submission...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    
    # Header
    headers = ['State', 'Constituency', 'Predicted Winner']
    ws.append(headers)
    
    # Format header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for idx, row in merged_df.iterrows():
        ws.append([row['state'], row['constituency'], row['predicted_winner']])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    
    # Freeze header
    ws.freeze_panes = 'A2'
    
    # Save
    excel_path = 'India_Predicts_2026_SUBMISSION.xlsx'
    wb.save(excel_path)
    logger.info(f"✓ Saved to {excel_path}")
    
    logger.info(f"\n{'='*70}")
    logger.info("✅ INTEGRATION COMPLETE")
    logger.info(f"{'='*70}\n")
    logger.info(f"Final submission ready: {excel_path}")


if __name__ == '__main__':
    result = integrate_team_candidates()
    
    if result is not None:
        logger.info("\n✅ Ready to submit!")
    else:
        logger.info("\n⚠ Please provide team candidate data file")
